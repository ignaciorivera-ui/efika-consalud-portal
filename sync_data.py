#!/usr/bin/env python3
"""
sync_data.py — EFIKA Portal Consalud
Descarga el Excel de SharePoint via Microsoft Graph API y genera data.js

Uso local:
  python sync_data.py                      # descarga y genera data.js
  python sync_data.py --push               # + git push a GitHub Pages
  python sync_data.py --local archivo.xlsx # usar Excel local (sin SharePoint)

Variables de entorno requeridas (o en archivo .env):
  AZURE_TENANT_ID      = bdabe570-1ae7-4cd4-8deb-7243d62c4520
  AZURE_CLIENT_ID      = (de tu Azure App Registration)
  AZURE_CLIENT_SECRET  = (de tu Azure App Registration)
"""

import sys, os, json, subprocess, tempfile
from datetime import datetime
from pathlib import Path

# ── CARGAR .env si existe ────────────────────────────────────────────────────
env_file = Path(__file__).parent / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

# ── CONFIG AZURE ──────────────────────────────────────────────────────────────
TENANT_ID     = os.environ.get("AZURE_TENANT_ID",     "bdabe570-1ae7-4cd4-8deb-7243d62c4520")
CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID",     "")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")

# Ruta del archivo en SharePoint (relativa al site EFIKA)
SP_FILE_PATH  = "EFIKA - Documentos/CLIENTES ACTUALES/CONSALUD/Control De Servicio/Control Consolidado Servicios/Control Consolidado Servicios.xlsx"
SP_SITE       = "efika2022.sharepoint.com:/sites/EFIKA:"

# ── KPIs BASE (se actualizan con los datos del Excel) ────────────────────────
KPI_BASE = {
    "totalAhorro":      1_234_112_357,
    "iniciConsultoria": "Junio 2022",
    "promMensual":      41.03,
    "factGestionadas":  1700,
    "privados": {
        "ahorro":       734_708_161,
        "pctAhorro":    43.97,
        "lineaBase":     63_512_363,
        "facturacion": 2_042_408_336,
        "cuotaActual":  36,
        "cuotaTotal":   36,
    },
    "moviles": {
        "ahorro":       174_558_177,
        "pctAhorro":    66.22,
        "lineaBase":      5_608_874,
        "facturacion":  111_026_179,
        "cuotaActual":  36,
        "cuotaTotal":   36,
    },
    "sms": {
        "ahorro":       112_255_502,
        "pctAhorro":    16.27,
        "lineaBase":    None,
        "facturacion":  334_952_933,
        "cuotaActual":  36,
        "cuotaTotal":   36,
    },
    "fija": {
        "ahorro":        49_720_126,
        "pctAhorro":    35.83,
        "lineaBase":      5_530_474,
        "facturacion":   45_796_690,
        "cuotaActual":  36,
        "cuotaTotal":   36,
    },
    "otras": {
        "totalPack":     16_909_163,
        "medical":       59_988_360,
        "microsoft":    132_077_360,
    },
}

# ── GRAPH API: OBTENER TOKEN ──────────────────────────────────────────────────
def get_token():
    try:
        import requests
    except ImportError:
        print("  ⚠  Instala requests: pip install requests")
        return None

    if not CLIENT_ID or not CLIENT_SECRET:
        print("  ⚠  AZURE_CLIENT_ID y AZURE_CLIENT_SECRET no configurados")
        return None

    r = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={
            "client_id":     CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "scope":         "https://graph.microsoft.com/.default",
            "grant_type":    "client_credentials",
        },
        timeout=20
    )
    if r.status_code != 200:
        print(f"  ✗ Error de autenticación: {r.text[:200]}")
        return None
    return r.json()["access_token"]

# ── GRAPH API: DESCARGAR EXCEL ────────────────────────────────────────────────
def download_excel_from_sharepoint():
    try:
        import requests
    except ImportError:
        return None

    print("  🔑 Autenticando con Microsoft Graph...")
    token = get_token()
    if not token:
        return None

    headers = {"Authorization": f"Bearer {token}"}

    # Obtener el archivo vía path relativo al site
    url = f"https://graph.microsoft.com/v1.0/sites/{SP_SITE}/drive/root:/{SP_FILE_PATH}:/content"
    print(f"  📥 Descargando Excel desde SharePoint...")
    r = requests.get(url, headers=headers, timeout=60)

    if r.status_code == 200:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(r.content)
        tmp.close()
        print(f"  ✓ Excel descargado ({len(r.content)//1024} KB)")
        return Path(tmp.name)
    else:
        print(f"  ✗ Error descargando Excel: {r.status_code} — {r.text[:200]}")
        return None

# ── LEER EXCEL ────────────────────────────────────────────────────────────────
def read_excel(path):
    try:
        import openpyxl
    except ImportError:
        print("  ⚠  Instala openpyxl: pip install openpyxl")
        return []

    print(f"  📊 Procesando: {Path(path).name}")
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        print(f"  ✗ Error abriendo Excel: {e}")
        return []

    # Intentar hojas conocidas
    ws = None
    for name in ["Resumen Ahorros", "Resumen", "Sheet1", wb.sheetnames[0]]:
        if name in wb.sheetnames:
            ws = wb[name]
            break

    if ws is None:
        print(f"  ✗ No se encontró hoja válida. Hojas disponibles: {wb.sheetnames}")
        return []

    print(f"  📋 Hoja: '{ws.title}'")
    rows = list(ws.iter_rows(values_only=True))
    monthly = []

    for r in rows[1:]:
        if not r or not r[0]:
            continue
        if not isinstance(r[0], datetime):
            continue
        monthly.append({
            "m":     r[0].strftime("%b %y"),
            "mv":    round((r[1] or 0) / 1e6, 2),
            "sms":   round((r[2] or 0) / 1e6, 2),
            "pv":    round((r[3] or 0) / 1e6, 2),
            "med":   round((r[4] or 0) / 1e6, 2),
            "ms":    round((r[5] or 0) / 1e6, 2),
            "fi":    round((r[7] or 0) / 1e6, 2) if len(r) > 7 else 0,
            "total": round((r[8] or 0) / 1e6, 2) if len(r) > 8 else 0,
        })

    print(f"  ✓ {len(monthly)} meses leídos")
    return monthly

# ── GENERAR data.js ───────────────────────────────────────────────────────────
def write_data_js(monthly, kpis, output_path):
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    kpis["updatedAt"] = ts
    kpis["mesesGestion"] = len(monthly) if monthly else 47

    if monthly:
        total_m = sum(r.get("total", 0) for r in monthly)
        if total_m > 1:  # sanity check — si el total parece real
            kpis["totalAhorro"] = int(total_m * 1e6)

    data = {"kpis": kpis, "monthly": monthly}
    js = (
        f"// Auto-generado por sync_data.py — {ts}\n"
        f"// Fuente: SharePoint EFIKA → Control Consolidado Servicios.xlsx\n"
        f"const PORTAL_DATA = {json.dumps(data, ensure_ascii=False, indent=2)};\n"
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(js)

    size_kb = output_path.stat().st_size // 1024
    print(f"  ✓ data.js generado — {len(monthly)} meses — {size_kb} KB — {ts}")

# ── GIT PUSH ──────────────────────────────────────────────────────────────────
def git_push(folder):
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    cmds = [
        ["git", "-C", str(folder), "add", "data.js"],
        ["git", "-C", str(folder), "commit", "-m", f"Auto-sync datos SharePoint {ts}"],
        ["git", "-C", str(folder), "push"],
    ]
    for cmd in cmds:
        r = subprocess.run(cmd, capture_output=True, text=True)
        label = " ".join(cmd[2:4])
        if r.returncode != 0:
            out = r.stderr.strip() or r.stdout.strip()
            if "nothing to commit" in out:
                print(f"  ℹ  {label}: sin cambios, no hay push necesario")
            else:
                print(f"  ⚠  {label}: {out[:120]}")
        else:
            print(f"  ✓  {label}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    do_push   = "--push" in args
    local_arg = next((a for a in args if not a.startswith("--")), None)

    print("\n╔══════════════════════════════════════════════════╗")
    print("║  EFIKA — Sync Portal Consalud (Microsoft Graph)  ║")
    print("╚══════════════════════════════════════════════════╝\n")

    folder  = Path(__file__).parent
    output  = folder / "data.js"
    tmp_xl  = None

    # --- Origen del Excel ---
    if local_arg and Path(local_arg).exists():
        xl_path = Path(local_arg)
        print(f"  📂 Modo local: {xl_path.name}")
    else:
        xl_path = None
        # Buscar en carpeta del script
        for pat in ["Control Consolidado*.xlsx", "*.xlsx"]:
            found = list(folder.glob(pat))
            if found:
                xl_path = found[0]
                print(f"  📂 Excel local encontrado: {xl_path.name}")
                break

        if xl_path is None:
            # Descargar desde SharePoint
            xl_path = download_excel_from_sharepoint()
            tmp_xl = xl_path  # marcar para borrar al final

    if xl_path is None:
        print("  ✗ No hay Excel disponible. Genera el data.js con datos existentes.")
        write_data_js([], dict(KPI_BASE), output)
    else:
        monthly = read_excel(xl_path)
        write_data_js(monthly, dict(KPI_BASE), output)

    # Limpiar temporal
    if tmp_xl and tmp_xl.exists():
        tmp_xl.unlink()

    if do_push:
        print("\n  📤 Subiendo a GitHub Pages...")
        git_push(folder)
        print("\n  🌐 Portal se actualiza en ~60 segundos en GitHub Pages")
    else:
        print("\n  💡 Usa --push para subir automáticamente a GitHub Pages")

    print("\n  ✅ Listo\n")

if __name__ == "__main__":
    main()
